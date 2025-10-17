import csv
import json
import os
import sys
from decimal import Decimal
from pathlib import Path

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook

sys.path.append(str(Path(__file__).resolve().parents[1]))

from backend.application import get_workspace_service, reset_workspace_state


@pytest.fixture(autouse=True)
def reset_state():
    reset_workspace_state()
    yield
    reset_workspace_state()


@pytest.fixture()
def client(tmp_path, monkeypatch):
    monkeypatch.setenv("WORKSPACES_ROOT", str(tmp_path))
    from importlib import reload
    from backend.core import workspaces

    reload(workspaces)
    from backend.app import create_app

    app = create_app()
    with TestClient(app) as test_client:
        yield test_client


def _write_csv(tmp_path: Path, filename: str, rows: list[dict]) -> Path:
    if not rows:
        raise ValueError("rows must not be empty")

    fieldnames: list[str] = []
    for row in rows:
        for key in row.keys():
            if key not in fieldnames:
                fieldnames.append(key)

    path = tmp_path / filename
    with path.open("w", encoding="utf-8", newline="") as fp:
        writer = csv.DictWriter(fp, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow(row)
    return path


def test_end_to_end_workflow(client, tmp_path):
    # 1. create workspace
    response = client.post("/api/workspaces", json={"month": "2025-01"})
    assert response.status_code == 200
    ws_id = response.json()["ws_id"]

    # 2. upload fact csv
    fact_rows = [
        {
            "employee_name": "张三",
            "period_month": "2025-01",
            "metric_code": "HOUR_TOTAL",
            "metric_value": 160,
            "unit": "hour",
            "confidence": 0.95,
        },
        {
            "employee_name": "张三",
            "period_month": "2025-01",
            "metric_code": "AMOUNT_BASE",
            "metric_value": 10000,
            "unit": "currency",
        },
        {
            "employee_name": "张三",
            "period_month": "2025-01",
            "metric_code": "AMOUNT_ALLOW",
            "metric_value": 500,
            "unit": "currency",
        },
    ]
    fact_path = _write_csv(tmp_path, "facts.csv", fact_rows)
    with fact_path.open("rb") as fp:
        response = client.post(
            f"/api/workspaces/{ws_id}/upload",
            files={"file": ("facts.csv", fp, "text/csv")},
        )
    assert response.status_code == 200
    job_id = response.json()["job_id"]

    # Verify job is completed
    response = client.get(f"/api/workspaces/{ws_id}/files")
    data = response.json()
    job = next(item for item in data["files"] if item["job_id"] == job_id)
    assert job["status"] == "completed"

    # 3. upload policy csv
    policy_rows = [
        {
            "employee_name_norm": "张三",
            "period_month": "2025-01",
            "mode": "SALARIED",
            "base_amount": 10000,
            "ot_weekday_rate": 50,
            "ot_weekend_rate": 80,
            "social_security_json": {"employee": 0.1},
        }
    ]
    policy_path = _write_csv(tmp_path, "policy.csv", policy_rows)
    with policy_path.open("rb") as fp:
        response = client.post(
            f"/api/workspaces/{ws_id}/upload",
            files={"file": ("policy.csv", fp, "text/csv")},
        )
    assert response.status_code == 200

    # 4. query fact and policy snapshots
    response = client.get(f"/api/workspaces/{ws_id}/fact")
    facts = response.json()["items"]
    assert len(facts) >= 3

    response = client.get(f"/api/workspaces/{ws_id}/policy")
    policies = response.json()["items"]
    assert policies and policies[0]["employee_name_norm"] == "张三"

    # 5. trigger calculation
    response = client.post(
        f"/api/workspaces/{ws_id}/calc",
        json={"period": "2025-01", "selected": ["张三"]},
    )
    assert response.status_code == 200
    assert response.json()["items"]

    # 6. fetch results
    response = client.get(f"/api/workspaces/{ws_id}/results", params={"period": "2025-01"})
    items = response.json()["items"]
    assert items and items[0]["employee_name_norm"] == "张三"


def test_unstructured_document_pipeline(client, tmp_path):
    response = client.post("/api/workspaces", json={"month": "2025-05"})
    assert response.status_code == 200
    ws_id = response.json()["ws_id"]

    image_path = tmp_path / "receipt.png"
    image_path.write_bytes(b"\x89PNG\r\n\x1a\n")

    with image_path.open("rb") as fp:
        upload = client.post(
            f"/api/workspaces/{ws_id}/upload",
            files={"file": ("receipt.png", fp, "image/png")},
        )
    assert upload.status_code == 200

    service = get_workspace_service()
    documents = service.list_documents(ws_id)
    assert len(documents) == 1
    document = documents[0]
    assert document["source_file"] == "receipt.png"
    assert document["schema"] == "image_document"
    assert document["requires_ocr"] is True
    assert document["ocr_metadata"]["provider"] == "noop"

    ocr_json = tmp_path / ws_id / "ocr" / "receipt.json"
    assert ocr_json.exists()
    payload = json.loads(ocr_json.read_text(encoding="utf-8"))
    assert payload["schema"] == "image_document"


def test_excel_templates_pipeline(client, tmp_path):
    response = client.post("/api/workspaces", json={"month": "2025-02"})
    assert response.status_code == 200
    ws_id = response.json()["ws_id"]

    # create timesheet aggregate workbook
    agg_wb = Workbook()
    agg_ws = agg_wb.active
    agg_ws.title = "汇总"
    agg_ws.append(["序号", "部门", "姓名", "工作日标准工时", "工作日加班工时", "周末节假日打卡工时", "当月工时（已公式加和）", "确认工时"])
    agg_ws.append([1, "财务部", "李四", 160, 12, 6, 178, 178])
    agg_path = tmp_path / "timesheet.xlsx"
    agg_wb.save(agg_path)

    with agg_path.open("rb") as fp:
        upload = client.post(f"/api/workspaces/{ws_id}/upload", files={"file": ("timesheet.xlsx", fp, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
    assert upload.status_code == 200

    # create policy workbook
    policy_wb = Workbook()
    policy_ws = policy_wb.active
    policy_ws.title = "薪资"
    policy_ws.append(["姓名", "模式", "基本工资", "工作日加班费率", "周末加班费率", "社保个人比例"])
    policy_ws.append(["李四", "SALARIED", 12000, 50, 80, 0.08])
    policy_path = tmp_path / "policy.xlsx"
    policy_wb.save(policy_path)

    with policy_path.open("rb") as fp:
        upload = client.post(f"/api/workspaces/{ws_id}/upload", files={"file": ("policy.xlsx", fp, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
    assert upload.status_code == 200

    facts = client.get(f"/api/workspaces/{ws_id}/fact").json()["items"]
    assert any(item["employee_name_norm"] == "李四" and item["metric_code"] == "HOUR_TOTAL" for item in facts)

    policies = client.get(f"/api/workspaces/{ws_id}/policy").json()["items"]
    assert any(item["employee_name_norm"] == "李四" for item in policies)

    calc = client.post(f"/api/workspaces/{ws_id}/calc", json={"period": "2025-02", "selected": ["李四"]})
    assert calc.status_code == 200
    assert calc.json()["items"]


def test_workspace_progress_tracking(client, tmp_path):
    response = client.post("/api/workspaces", json={"month": "2025-09"})
    assert response.status_code == 200
    ws_id = response.json()["ws_id"]

    list_response = client.get("/api/workspaces")
    assert list_response.status_code == 200
    assert any(item["ws_id"] == ws_id for item in list_response.json()["items"])

    progress = client.get(f"/api/workspaces/{ws_id}/progress").json()
    step_map = {step["id"]: step for step in progress["steps"]}
    assert step_map["workspace_setup"]["status"] == "completed"
    assert step_map["upload_timesheets"]["status"] == "pending"

    # 上传 timesheet_personal 模板
    timesheet_wb = Workbook()
    timesheet_ws = timesheet_wb.active
    timesheet_ws.title = "个人工时"
    timesheet_ws["A1"] = "姓名"
    timesheet_ws["B1"] = "张三"
    timesheet_ws["A2"] = "月份"
    timesheet_ws["B2"] = "2025-09"
    timesheet_ws.append([])
    timesheet_ws.append(["日期", "标准工时", "加班工时", "周末节假日打卡工时", "总工时"])
    timesheet_ws.append(["2025-09-01", 8, 2, 0, 10])
    timesheet_ws.append(["2025-09-02", 8, 1, 0, 9])
    timesheet_path = tmp_path / "personal.xlsx"
    timesheet_wb.save(timesheet_path)

    with timesheet_path.open("rb") as fp:
        upload_ts = client.post(
            f"/api/workspaces/{ws_id}/upload",
            files={"file": ("personal.xlsx", fp, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
    assert upload_ts.status_code == 200

    progress = client.get(f"/api/workspaces/{ws_id}/progress").json()
    step_map = {step["id"]: step for step in progress["steps"]}
    assert step_map["upload_timesheets"]["status"] in {"in_progress", "completed"}
    ts_requirements = step_map["upload_timesheets"]["requirements"]
    assert any(req["id"] == "timesheet_detail" and req["status"] == "completed" for req in ts_requirements)

    # 上传 roster_sheet 模板
    roster_path = _write_csv(
        tmp_path,
        "roster.csv",
        [
            {
                "姓名": "张三",
                "个人比例": 0.08,
                "公司比例": 0.1,
                "最低基数": 5000,
                "最高基数": 15000,
                "入职日期": "2023-01-01",
                "离职日期": "",
                "月份": "2025-09",
            }
        ],
    )

    with roster_path.open("rb") as fp:
        upload_roster = client.post(
            f"/api/workspaces/{ws_id}/upload",
            files={"file": ("roster.csv", fp, "text/csv")},
        )
    assert upload_roster.status_code == 200

    # 上传 policy_sheet 模板
    policy_wb = Workbook()
    policy_ws = policy_wb.active
    policy_ws.title = "薪资"
    policy_ws.append(["姓名", "模式", "基本工资", "工作日加班费率", "周末加班费率", "社保个人比例"])
    policy_ws.append(["张三", "SALARIED", 12000, 50, 80, 0.08])
    policy_path = tmp_path / "policy.xlsx"
    policy_wb.save(policy_path)

    with policy_path.open("rb") as fp:
        upload_policy = client.post(
            f"/api/workspaces/{ws_id}/upload",
            files={"file": ("policy.xlsx", fp, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
    assert upload_policy.status_code == 200

    progress = client.get(f"/api/workspaces/{ws_id}/progress").json()
    step_map = {step["id"]: step for step in progress["steps"]}
    assert step_map["upload_policy"]["status"] == "completed"
    policy_requirements = step_map["upload_policy"]["requirements"]
    assert all(req["status"] == "completed" or req["optional"] for req in policy_requirements)

    # 标记审查完成
    checkpoint = client.post(
        f"/api/workspaces/{ws_id}/progress/checkpoints",
        json={"step": "review_data", "status": "completed"},
    )
    assert checkpoint.status_code == 200
    progress = checkpoint.json()["progress"]
    review_step = next(step for step in progress["steps"] if step["id"] == "review_data")
    assert review_step["status"] == "completed"

    # 触发计算并检查最终状态
    calc = client.post(
        f"/api/workspaces/{ws_id}/calc",
        json={"period": "2025-09", "selected": ["张三"]},
    )
    assert calc.status_code == 200
    progress = client.get(f"/api/workspaces/{ws_id}/progress").json()
    payroll_step = next(step for step in progress["steps"] if step["id"] == "run_payroll")
    assert payroll_step["status"] == "completed"


def test_official_templates_pipeline(client):
    response = client.post("/api/workspaces", json={"month": "2025-10"})
    assert response.status_code == 200
    ws_id = response.json()["ws_id"]

    templates_dir = Path(__file__).resolve().parents[1] / "samples" / "templates"
    uploads = [
        ("timesheet_personal_template.csv", "text/csv"),
        ("timesheet_aggregate_template.csv", "text/csv"),
        ("roster_sheet_template.csv", "text/csv"),
        ("policy_sheet_template.csv", "text/csv"),
    ]

    for filename, content_type in uploads:
        file_path = templates_dir / filename
        with file_path.open("rb") as fp:
            upload = client.post(
                f"/api/workspaces/{ws_id}/upload",
                files={"file": (filename, fp, content_type)},
            )
        assert upload.status_code == 200

    calc = client.post(
        f"/api/workspaces/{ws_id}/calc",
        json={"period": "2025-10"},
    )
    assert calc.status_code == 200
    items = calc.json()["items"]
    assert items

    net_values = [Decimal(str(item["net_pay"])) for item in items]
    base_values = [Decimal(str(item["base_pay"])) for item in items]

    assert all(value != Decimal("0") for value in net_values)
    assert all(value > Decimal("0") for value in base_values)


def test_policy_roster_merge_preserves_base_amount(client, tmp_path):
    response = client.post("/api/workspaces", json={"month": "2025-03"})
    assert response.status_code == 200
    ws_id = response.json()["ws_id"]

    fact_rows = [
        {
            "employee_name": "张三",
            "period_month": "2025-03",
            "metric_code": "HOUR_TOTAL",
            "metric_value": 160,
            "unit": "hour",
        }
    ]
    fact_path = _write_csv(tmp_path, "facts_merge.csv", fact_rows)
    with fact_path.open("rb") as fp:
        upload_fact = client.post(
            f"/api/workspaces/{ws_id}/upload",
            files={"file": ("facts_merge.csv", fp, "text/csv")},
        )
    assert upload_fact.status_code == 200

    policy_rows = [
        {
            "employee_name_norm": "张三",
            "period_month": "2025-03",
            "mode": "SALARIED",
            "base_amount": 10000,
            "social_security_json": {"employee": 0.0},
        }
    ]
    policy_path = _write_csv(tmp_path, "policy_merge.csv", policy_rows)
    with policy_path.open("rb") as fp:
        upload_policy = client.post(
            f"/api/workspaces/{ws_id}/upload",
            files={"file": ("policy_merge.csv", fp, "text/csv")},
        )
    assert upload_policy.status_code == 200

    roster_rows = [
        {
            "姓名": "张三",
            "个人比例": 0.08,
            "公司比例": 0.1,
            "最低基数": 5000,
            "最高基数": 20000,
            "月份": "2025-03",
        }
    ]
    roster_path = _write_csv(tmp_path, "roster_merge.csv", roster_rows)
    with roster_path.open("rb") as fp:
        upload_roster = client.post(
            f"/api/workspaces/{ws_id}/upload",
            files={"file": ("roster_merge.csv", fp, "text/csv")},
        )
    assert upload_roster.status_code == 200

    calc = client.post(
        f"/api/workspaces/{ws_id}/calc",
        json={"period": "2025-03", "selected": ["张三"]},
    )
    assert calc.status_code == 200
    item = calc.json()["items"][0]
    assert Decimal(str(item["base_pay"])) > Decimal("0")
    assert Decimal(str(item["social_security_personal"])) == Decimal("800.00")


def test_timesheet_with_metadata_header_rows(client, tmp_path):
    response = client.post("/api/workspaces", json={"month": "2025-04"})
    assert response.status_code == 200
    ws_id = response.json()["ws_id"]

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "月度工时"
    sheet.append(["填报单位", "卓迈科技", None, None, None, None, None, None])
    sheet.append(["月份", "2025-04", None, None, None, None, None, None])
    sheet.append([None, None, None, None, None, None, None, None])
    sheet.append(
        [
            "序号",
            "部门",
            "姓名",
            "工作日标准工时",
            "工作日加班工时",
            "周末节假日打卡工时",
            "当月工时（已公式加和）",
            "确认工时",
        ]
    )
    sheet.append([1, "生产部", "赵六", 150, 20, 8, 178, 178])
    sheet.append([2, "生产部", "钱七", 160, 18, 6, 184, 184])

    path = tmp_path / "metadata_timesheet.xlsx"
    workbook.save(path)

    with path.open("rb") as fp:
        upload = client.post(
            f"/api/workspaces/{ws_id}/upload",
            files={
                "file": (
                    "metadata_timesheet.xlsx",
                    fp,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
    assert upload.status_code == 200

    facts = client.get(f"/api/workspaces/{ws_id}/fact").json()["items"]
    assert any(item["employee_name_norm"] == "赵六" and item["metric_code"] == "HOUR_TOTAL" for item in facts)
    assert any(item["employee_name_norm"] == "钱七" and item["metric_code"] == "HOUR_STD" for item in facts)


def test_policy_ingestion_case_insensitive_columns(client, tmp_path):
    response = client.post("/api/workspaces", json={"month": "2025-05"})
    assert response.status_code == 200
    ws_id = response.json()["ws_id"]

    policy_rows = [
        {
            "Employee_Name_Norm": "王五",
            "Period_Month": "2025-05",
            "Mode": "salaried",
            "Base_Amount": 12500,
            "Ot_Weekday_Rate": 45,
            "Social_Security_Json": {"employee": 0.08},
        }
    ]
    policy_path = _write_csv(tmp_path, "policy_case.csv", policy_rows)
    with policy_path.open("rb") as fp:
        upload = client.post(
            f"/api/workspaces/{ws_id}/upload",
            files={"file": ("policy_case.csv", fp, "text/csv")},
        )
    assert upload.status_code == 200

    response = client.get(f"/api/workspaces/{ws_id}/policy")
    items = response.json()["items"]
    assert items
    policy = next(item for item in items if item["employee_name_norm"] == "王五")

    assert policy["mode"] == "SALARIED"
    assert Decimal(str(policy["base_amount"])) == Decimal("12500")
    assert Decimal(str(policy["ot_weekday_rate"])) == Decimal("45")
    assert policy["ot_weekend_rate"] is None
    assert policy["social_security_json"] == {"employee": 0.08}


def test_period_month_normalisation_for_localised_strings(client, tmp_path):
    response = client.post("/api/workspaces", json={"month": "2025-08"})
    assert response.status_code == 200
    ws_id = response.json()["ws_id"]

    fact_rows = [
        {
            "employee_name": "李雷",
            "period_month": "八月-Aug",
            "metric_code": "HOUR_TOTAL",
            "metric_value": 172,
            "unit": "hour",
        }
    ]
    fact_path = _write_csv(tmp_path, "localized_fact.csv", fact_rows)
    with fact_path.open("rb") as fp:
        upload_fact = client.post(
            f"/api/workspaces/{ws_id}/upload",
            files={"file": ("localized_fact.csv", fp, "text/csv")},
        )
    assert upload_fact.status_code == 200

    policy_rows = [
        {
            "employee_name_norm": "李雷",
            "period_month": "八月-Aug",
            "mode": "SALARIED",
            "base_amount": 9800,
        }
    ]
    policy_path = _write_csv(tmp_path, "localized_policy.csv", policy_rows)
    with policy_path.open("rb") as fp:
        upload_policy = client.post(
            f"/api/workspaces/{ws_id}/upload",
            files={"file": ("localized_policy.csv", fp, "text/csv")},
        )
    assert upload_policy.status_code == 200

    facts = client.get(f"/api/workspaces/{ws_id}/fact").json()["items"]
    fact = next(item for item in facts if item["employee_name_norm"] == "李雷")
    assert fact["period_month"] == "2025-08"

    policies = client.get(f"/api/workspaces/{ws_id}/policy").json()["items"]
    policy = next(item for item in policies if item["employee_name_norm"] == "李雷")
    assert policy["period_month"] == "2025-08"


def test_excel_heuristic_fallback(client, tmp_path):
    response = client.post("/api/workspaces", json={"month": "2025-03"})
    assert response.status_code == 200
    ws_id = response.json()["ws_id"]

    # workbook with unconventional headers so the template parser emits no rows
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "原始工时"
    sheet.append(["员工姓名", "本月正常工时", "平时加班时长", "周末加班总时长", "核定工时"])
    sheet.append(["王五", 150, 10, 8, 168])
    timesheet_path = tmp_path / "heuristic_timesheet.xlsx"
    workbook.save(timesheet_path)

    with timesheet_path.open("rb") as fp:
        upload = client.post(
            f"/api/workspaces/{ws_id}/upload",
            files={"file": ("heuristic_timesheet.xlsx", fp, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
    assert upload.status_code == 200

    policy_wb = Workbook()
    policy_ws = policy_wb.active
    policy_ws.title = "薪资"
    policy_ws.append(["姓名", "月薪", "平日加班费率", "周末加班费率", "社保个人比例", "餐补津贴"])
    policy_ws.append(["王五", 9000, 45, 70, 0.08, 150])
    policy_path = tmp_path / "heuristic_policy.xlsx"
    policy_wb.save(policy_path)

    with policy_path.open("rb") as fp:
        upload_policy = client.post(
            f"/api/workspaces/{ws_id}/upload",
            files={"file": ("heuristic_policy.xlsx", fp, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
    assert upload_policy.status_code == 200

    facts = client.get(f"/api/workspaces/{ws_id}/fact").json()["items"]
    assert any(item["employee_name_norm"] == "王五" and item["metric_code"] == "HOUR_CONFIRMED" for item in facts)

    policies = client.get(f"/api/workspaces/{ws_id}/policy").json()["items"]
    assert any(item["employee_name_norm"] == "王五" and item.get("base_amount") for item in policies)

    calc = client.post(
        f"/api/workspaces/{ws_id}/calc",
        json={"period": "2025-03", "selected": ["王五"]},
    )
    assert calc.status_code == 200
    assert calc.json()["items"]
