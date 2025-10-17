"""Parser for the 个人工时表."""

from __future__ import annotations

from dataclasses import dataclass
from decimal import Decimal
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook

from backend.core import name_normalize


@dataclass
class PersonalParseResult:
    facts: list[dict[str, Any]]


def _safe_decimal(value: Any) -> Decimal | None:
    if value is None or value == "":
        return None
    try:
        decimal_value = Decimal(str(value))
    except Exception:  # pragma: no cover - defensive
        return None
    if not decimal_value.is_finite():
        return None
    return decimal_value


def _find_metadata(ws, keyword: str) -> str | None:
    for row in ws.iter_rows(min_row=1, max_row=15, max_col=10):
        for cell in row:
            value = cell.value
            if value and keyword in str(value):
                neighbour = ws.cell(row=cell.row, column=cell.column + 1)
                if neighbour.value is None:
                    continue
                return str(neighbour.value).strip()
    return None


def _find_header_row(ws) -> int | None:
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        labels = [str(cell.value).strip() if cell.value is not None else "" for cell in row]
        if "日期" in labels and ("总工时" in labels or "工时" in labels):
            return row[0].row
    return None


def _normalise_columns(dataframe: pd.DataFrame) -> pd.DataFrame:
    renamed = {col: str(col).strip() for col in dataframe.columns}
    return dataframe.rename(columns=renamed)


def _find_column(dataframe: pd.DataFrame, keywords: list[str]) -> str | None:
    for keyword in keywords:
        for column in dataframe.columns:
            if keyword in str(column):
                return column
    return None


def _parse_csv(path: Path, ws_id: str, period: str | None = None) -> PersonalParseResult:
    dataframe = pd.read_csv(path)
    dataframe = dataframe.dropna(how="all")
    dataframe = _normalise_columns(dataframe)

    name_column = _find_column(dataframe, ["姓名", "员工", "name"])
    month_column = _find_column(dataframe, ["月份", "月度", "period"])

    employee_name = ""
    if name_column and not dataframe[name_column].dropna().empty:
        employee_name = str(dataframe[name_column].dropna().iloc[0]).strip()

    month = period or ws_id
    if month_column and not dataframe[month_column].dropna().empty:
        month = str(dataframe[month_column].dropna().iloc[0]).strip()

    metric_columns: dict[str, str] = {}
    for metric, keywords in {
        "HOUR_TOTAL": ["总工时"],
        "HOUR_STD": ["标准工时", "工作日标准工时"],
        "HOUR_OT_WD": ["加班工时", "工作日加班工时"],
        "HOUR_OT_WE": ["周末节假日打卡工时", "周末加班工时"],
    }.items():
        column = _find_column(dataframe, keywords)
        if column:
            metric_columns[metric] = column

    totals = {metric: Decimal("0") for metric in metric_columns}
    for metric_code, column in metric_columns.items():
        for value in dataframe[column].tolist():
            decimal = _safe_decimal(value)
            if decimal is None:
                continue
            totals[metric_code] += decimal

    facts: list[dict[str, Any]] = []
    if employee_name:
        for metric_code, total in totals.items():
            if total == 0:
                continue
            facts.append(
                {
                    "employee_name": employee_name,
                    "employee_name_norm": name_normalize.normalize(employee_name),
                    "period_month": month,
                    "metric_code": metric_code,
                    "metric_value": total,
                    "unit": "hour",
                    "metric_label": metric_columns.get(metric_code, "个人工时"),
                    "confidence": Decimal("0.8"),
                    "source_sheet": None,
                }
            )

    return PersonalParseResult(facts=facts)


def parse(path: Path, ws_id: str, sheet_name: str | None = None, period: str | None = None) -> PersonalParseResult:
    if path.suffix.lower() == ".csv":
        return _parse_csv(path, ws_id, period)

    workbook = load_workbook(path, data_only=True)
    worksheet = workbook[sheet_name] if sheet_name and sheet_name in workbook.sheetnames else workbook.active

    employee_name = _find_metadata(worksheet, "姓名") or ""
    month = _find_metadata(worksheet, "月份") or period or ws_id
    header_row = _find_header_row(worksheet)
    if not header_row:
        return PersonalParseResult(facts=[])

    headers = [str(cell.value).strip() if cell.value else "" for cell in worksheet[header_row]]
    indices = {header: idx for idx, header in enumerate(headers)}

    totals = {
        "HOUR_TOTAL": Decimal("0"),
        "HOUR_STD": Decimal("0"),
        "HOUR_OT_WD": Decimal("0"),
        "HOUR_OT_WE": Decimal("0"),
    }

    for row in worksheet.iter_rows(min_row=header_row + 1, values_only=True):
        if all(cell is None for cell in row):
            continue
        if row[0] in {"合计", "总计"}:
            continue

        if "总工时" in indices:
            value = _safe_decimal(row[indices["总工时"]])
            if value:
                totals["HOUR_TOTAL"] += value
        if "标准工时" in indices:
            value = _safe_decimal(row[indices["标准工时"]])
            if value:
                totals["HOUR_STD"] += value
        if "加班工时" in indices:
            value = _safe_decimal(row[indices["加班工时"]])
            if value:
                totals["HOUR_OT_WD"] += value
        if "周末节假日打卡工时" in indices:
            value = _safe_decimal(row[indices["周末节假日打卡工时"]])
            if value:
                totals["HOUR_OT_WE"] += value

    facts: list[dict[str, Any]] = []
    if employee_name:
        for metric, total in totals.items():
            if total == 0:
                continue
            facts.append(
                {
                    "employee_name": employee_name,
                    "employee_name_norm": name_normalize.normalize(employee_name),
                    "period_month": month,
                    "metric_code": metric,
                    "metric_value": total,
                    "unit": "hour",
                    "metric_label": "个人工时汇总",
                    "confidence": Decimal("0.8"),
                    "source_sheet": worksheet.title,
                }
            )

    return PersonalParseResult(facts=facts)

